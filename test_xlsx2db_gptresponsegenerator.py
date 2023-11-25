import openpyxl
import sqlite3
import os
from dotenv import load_dotenv
from GPTResponseGenerator import GPTResponseGenerator

def read_excel_data(file_path, rows, columns):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    
    for row in range(1, rows + 1):
        row_data = []
        for col in range(1, columns + 1):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        data.append(row_data)
    
    return data, workbook, sheet

def store_in_excel(data, responses, output_file_path, start_row):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Copying original data
    for idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=idx, column=col_idx).value = value

    # Storing responses
    for idx, response in enumerate(responses, start=start_row):
        sheet.cell(row=idx, column=len(data[0]) + 1).value = response

    workbook.save(output_file_path)

def store_in_sqlite(data, responses, db_file):
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS responses
                 (input_data TEXT, ai_response TEXT)''')

    combined_data = [(str(d), r) for d, r in zip(data, responses)]
    c.executemany('INSERT INTO responses VALUES (?, ?)', combined_data)
    conn.commit()
    conn.close()

def main():
    load_dotenv()
    system_message = os.getenv("XLSX2DB_SYSTEM_MESSAGE", "Default system message")
    input_file_path = os.getenv("EXCEL_INPUT_FILE_PATH", "input_data.xlsx")
    output_file_path = os.getenv("EXCEL_OUTPUT_FILE_PATH", "output_data.xlsx")
    rows = int(os.getenv("NUM_ROWS", 5))
    columns = int(os.getenv("NUM_COLUMNS", 3))
    db_file = os.getenv("SQLITE_DB_FILE", "responses.db")
    response_start_row = int(os.getenv("RESPONSE_START_ROW", 1))

    gpt_generator = GPTResponseGenerator()
    
    data, _, _ = read_excel_data(input_file_path, rows, columns)
    responses = [gpt_generator.generate_response(system_message, ' '.join(map(str, row))) for row in data]

    store_in_excel(data, responses, output_file_path, response_start_row)
    store_in_sqlite(data, responses, db_file)

if __name__ == "__main__":
    main()